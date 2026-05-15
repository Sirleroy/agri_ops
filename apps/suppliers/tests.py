import datetime

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from apps.companies.models import Company
from apps.suppliers.models import Supplier, Farm, DeforestationCheck
from apps.users.models import CustomUser


class SupplierTenantIsolationTest(TestCase):

    def setUp(self):
        self.company1 = Company.objects.create(
            name='Company One', country='Nigeria', plan_tier='starter'
        )
        self.company2 = Company.objects.create(
            name='Company Two', country='Nigeria', plan_tier='starter'
        )
        self.supplier1 = Supplier.objects.create(
            company=self.company1, name='Supplier One', category='seeds'
        )
        self.supplier2 = Supplier.objects.create(
            company=self.company2, name='Supplier Two', category='fertilizer'
        )

    def test_supplier_scoped_to_company(self):
        company1_suppliers = Supplier.objects.filter(company=self.company1)
        self.assertEqual(company1_suppliers.count(), 1)
        self.assertEqual(company1_suppliers.first().name, 'Supplier One')

    def test_no_cross_tenant_leakage(self):
        company1_suppliers = Supplier.objects.filter(company=self.company1)
        self.assertNotIn(self.supplier2, company1_suppliers)


_POLYGON = {
    'type': 'Polygon',
    'coordinates': [[
        [8.50, 9.00], [8.51, 9.00], [8.51, 9.01], [8.50, 9.01], [8.50, 9.00],
    ]],
}


class FarmComplianceTest(TestCase):
    """
    Compliance Readiness is evidence-backed: a sign-off (`is_eudr_verified`)
    only counts as `compliant` when a clear, current, non-stale deforestation
    check sits behind it. These tests guard that gate and the readiness/
    disqualification lifecycle the deforestation engine drives.
    """

    def setUp(self):
        self.company = Company.objects.create(
            name='Test Company', country='Nigeria', plan_tier='starter'
        )
        self.supplier = Supplier.objects.create(
            company=self.company, name='Test Supplier', category='seeds'
        )
        self.farm = Farm.objects.create(
            company=self.company,
            supplier=self.supplier,
            name='Test Farm',
            country='Nigeria',
            commodity='Soy',
            geolocation=_POLYGON,
            deforestation_risk_status='low',
            is_eudr_verified=False,
        )

    def _make_check(self, risk_status='clear', stale=False):
        return DeforestationCheck.objects.create(
            farm=self.farm,
            company=self.company,
            risk_status=risk_status,
            engine_status='complete',
            geometry_hash_at_assessment=(
                'stale-hash-0000' if stale else self.farm.geometry_hash
            ),
            assessed_at=timezone.now(),
        )

    def _sign_off(self):
        self.farm.is_eudr_verified = True
        self.farm.verified_date = datetime.date.today()
        self.farm.save()
        self.farm.refresh_from_db()

    # ── compliance_status ────────────────────────────────────────────────────

    def test_unverified_farm_is_pending(self):
        self.assertEqual(self.farm.compliance_status, 'pending')

    def test_verified_with_clear_current_check_is_compliant(self):
        self._make_check('clear')
        self._sign_off()
        self.assertEqual(self.farm.compliance_status, 'compliant')

    def test_verified_without_any_check_is_not_compliant(self):
        # The credibility hole this work closes: a sign-off with no evidence.
        self._sign_off()
        self.assertEqual(self.farm.compliance_status, 'pending')

    def test_verified_with_stale_check_is_not_compliant(self):
        self._make_check('clear', stale=True)
        self._sign_off()
        self.assertEqual(self.farm.compliance_status, 'pending')

    def test_flagged_check_disqualifies_farm(self):
        self._make_check('flagged')
        self.assertTrue(self.farm.is_disqualified)
        self.assertEqual(self.farm.compliance_status, 'disqualified')

    def test_expired_signoff_is_expired(self):
        self._make_check('clear')
        self.farm.is_eudr_verified = True
        self.farm.verification_expiry = datetime.date.today() - datetime.timedelta(days=1)
        self.farm.save()
        self.assertEqual(self.farm.compliance_status, 'expired')

    # ── disqualification: engine + manual override ───────────────────────────

    def test_manual_override_true_disqualifies(self):
        self.farm.land_cleared_after_cutoff = True
        self.assertTrue(self.farm.is_disqualified)

    def test_manual_override_false_beats_flagged_engine_check(self):
        self._make_check('flagged')
        self.farm.land_cleared_after_cutoff = False
        self.assertFalse(self.farm.is_disqualified)

    def test_override_false_clears_the_flagged_readiness_blocker(self):
        # An override to "not disqualified" must also clear the flagged blocker,
        # otherwise the audited override path can't actually reach sign-off.
        self._make_check('flagged')
        self.farm.land_cleared_after_cutoff = False
        self.assertEqual(self.farm.readiness_blockers, [])
        self.assertEqual(self.farm.readiness_state, 'awaiting_signoff')

    def test_no_override_defers_to_engine(self):
        self.assertIsNone(self.farm.land_cleared_after_cutoff)
        self._make_check('clear')
        self.assertFalse(self.farm.is_disqualified)

    def test_str_is_null_safe_without_supplier(self):
        # supplier is nullable; str(farm) is reached from the audit path.
        farm = Farm.objects.create(
            company=self.company, name='Supplierless Farm',
            country='Nigeria', commodity='Soy',
        )
        self.assertEqual(str(farm), 'Supplierless Farm')

    # ── readiness_state lifecycle ────────────────────────────────────────────

    def test_readiness_not_ready_without_check(self):
        self.assertEqual(self.farm.readiness_state, 'not_ready')
        self.assertTrue(self.farm.readiness_blockers)

    def test_readiness_awaiting_signoff_with_clear_check(self):
        self._make_check('clear')
        self.assertEqual(self.farm.readiness_state, 'awaiting_signoff')
        self.assertEqual(self.farm.readiness_blockers, [])

    def test_readiness_ready_after_signoff(self):
        self._make_check('clear')
        self._sign_off()
        self.assertEqual(self.farm.readiness_state, 'ready')

    def test_readiness_disqualified_with_flagged_check(self):
        self._make_check('flagged')
        self.assertEqual(self.farm.readiness_state, 'disqualified')

    # ── auto-invalidation ────────────────────────────────────────────────────

    def test_geometry_change_voids_signoff(self):
        self._make_check('clear')
        self._sign_off()
        self.assertTrue(self.farm.is_eudr_verified)
        # Edit the polygon — the sign-off no longer matches the boundary on file.
        moved = {
            'type': 'Polygon',
            'coordinates': [[
                [8.60, 9.10], [8.61, 9.10], [8.61, 9.11], [8.60, 9.11], [8.60, 9.10],
            ]],
        }
        self.farm.geolocation = moved
        self.farm.save()
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)
        self.assertIsNone(self.farm.verified_date)

    def test_flagged_recheck_voids_signoff(self):
        from unittest.mock import patch
        from apps.suppliers import deforestation_engine

        self._make_check('clear')
        self._sign_off()
        self.assertTrue(self.farm.is_eudr_verified)

        # A fresh engine run that comes back flagged must withdraw the sign-off.
        flagged_result = {
            'pixels': 120, 'loss_pixels': 6, 'loss_area_ha': 0.48,
            'loss_years': [21], 'risk_status': 'flagged', 'error': None,
        }
        with patch.object(deforestation_engine, 'intersect_farm', return_value=flagged_result):
            deforestation_engine.run_check(self.farm)

        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)
        self.assertIsNone(self.farm.verified_date)
        self.assertEqual(self.farm.compliance_status, 'disqualified')

        # The engine-triggered sign-off withdrawal must land in the audit log,
        # the same as a manual withdrawal would.
        from apps.audit.models import AuditLog
        log = (
            AuditLog.objects
            .filter(company=self.company, model_name='Farm', object_id=self.farm.pk)
            .order_by('-timestamp')
            .first()
        )
        self.assertIsNotNone(log)
        self.assertEqual(
            log.changes.get('compliance_readiness'), 'sign-off withdrawn — automated'
        )


class ComplianceReadinessSignoffTests(TestCase):
    """
    Compliance Readiness sign-off is an evidence-gated, manager-only,
    audit-logged action — not a free checkbox. These tests guard the evidence
    gate, the permission boundary, and tenant isolation.
    """

    def setUp(self):
        self.company = Company.objects.create(
            name='Sign-off Co', country='Nigeria', plan_tier='starter'
        )
        self.other_company = Company.objects.create(
            name='Other Co', country='Nigeria', plan_tier='starter'
        )
        self.supplier = Supplier.objects.create(
            company=self.company, name='Sign-off Supplier', category='cooperative'
        )
        self.manager = CustomUser.objects.create_user(
            username='signoff_mgr', password='testpass',
            company=self.company, system_role='manager',
        )
        self.staff = CustomUser.objects.create_user(
            username='signoff_staff', password='testpass',
            company=self.company, system_role='staff',
        )
        self.farm = Farm.objects.create(
            company=self.company, supplier=self.supplier, name='Sign-off Farm',
            country='Nigeria', commodity='Soy', geolocation=_POLYGON,
            deforestation_risk_status='low',
        )

    def _clear_check(self, farm=None):
        farm = farm or self.farm
        return DeforestationCheck.objects.create(
            farm=farm, company=farm.company, risk_status='clear',
            engine_status='complete',
            geometry_hash_at_assessment=farm.geometry_hash,
            assessed_at=timezone.now(),
        )

    def _confirm_url(self, farm=None):
        return reverse('suppliers:farm_confirm_readiness', kwargs={'pk': (farm or self.farm).pk})

    def _withdraw_url(self, farm=None):
        return reverse('suppliers:farm_withdraw_readiness', kwargs={'pk': (farm or self.farm).pk})

    def test_manager_can_sign_off_evidenced_farm(self):
        self._clear_check()
        self.client.force_login(self.manager)
        r = self.client.post(self._confirm_url())
        self.assertEqual(r.status_code, 302)
        self.farm.refresh_from_db()
        self.assertTrue(self.farm.is_eudr_verified)
        self.assertEqual(self.farm.verified_by, self.manager)
        self.assertEqual(self.farm.verified_date, datetime.date.today())
        self.assertEqual(
            self.farm.verification_expiry,
            datetime.date.today() + datetime.timedelta(days=365),
        )

    def test_signoff_rejected_without_evidence(self):
        # Farm has a polygon but no deforestation check — readiness has blockers.
        self.client.force_login(self.manager)
        r = self.client.post(self._confirm_url())
        self.assertEqual(r.status_code, 302)
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)

    def test_signoff_rejected_when_disqualified(self):
        DeforestationCheck.objects.create(
            farm=self.farm, company=self.company, risk_status='flagged',
            engine_status='complete',
            geometry_hash_at_assessment=self.farm.geometry_hash,
            assessed_at=timezone.now(),
        )
        self.client.force_login(self.manager)
        r = self.client.post(self._confirm_url())
        self.assertEqual(r.status_code, 302)
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)

    def test_staff_cannot_sign_off(self):
        self._clear_check()
        self.client.force_login(self.staff)
        r = self.client.post(self._confirm_url())
        self.assertEqual(r.status_code, 403)
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)

    def test_signoff_cross_tenant_is_404(self):
        other_supplier = Supplier.objects.create(
            company=self.other_company, name='Other Supplier', category='cooperative'
        )
        other_farm = Farm.objects.create(
            company=self.other_company, supplier=other_supplier, name='Other Farm',
            country='Nigeria', commodity='Soy', geolocation=_POLYGON,
        )
        self._clear_check(other_farm)
        self.client.force_login(self.manager)
        r = self.client.post(self._confirm_url(other_farm))
        self.assertEqual(r.status_code, 404)
        other_farm.refresh_from_db()
        self.assertFalse(other_farm.is_eudr_verified)

    def test_manager_can_withdraw_signoff(self):
        self._clear_check()
        self.client.force_login(self.manager)
        self.client.post(self._confirm_url())
        self.farm.refresh_from_db()
        self.assertTrue(self.farm.is_eudr_verified)
        r = self.client.post(self._withdraw_url())
        self.assertEqual(r.status_code, 302)
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)
        self.assertIsNone(self.farm.verified_by)

    def test_signoff_is_audit_logged(self):
        from apps.audit.models import AuditLog
        self._clear_check()
        self.client.force_login(self.manager)
        self.client.post(self._confirm_url())
        log = (
            AuditLog.objects
            .filter(company=self.company, model_name='Farm', object_id=self.farm.pk)
            .order_by('-timestamp')
            .first()
        )
        self.assertIsNotNone(log)
        self.assertEqual(log.changes.get('compliance_readiness'), 'signed off')

    def test_run_deforestation_check_is_audit_logged(self):
        """A manual 'Run Check' is a user action on a tenant record — it must
        appear in the central audit log, not only as a DeforestationCheck row."""
        from unittest.mock import patch
        from apps.suppliers import deforestation_engine
        from apps.audit.models import AuditLog

        clear_result = {
            'pixels': 200, 'loss_pixels': 0, 'loss_area_ha': 0,
            'loss_years': [], 'risk_status': 'clear', 'error': None,
        }
        self.client.force_login(self.staff)
        with patch.object(deforestation_engine, 'intersect_farm', return_value=clear_result):
            r = self.client.post(
                reverse('suppliers:run_deforestation_check', kwargs={'pk': self.farm.pk})
            )
        self.assertEqual(r.status_code, 302)
        log = (
            AuditLog.objects
            .filter(company=self.company, model_name='Farm', object_id=self.farm.pk)
            .order_by('-timestamp')
            .first()
        )
        self.assertIsNotNone(log)
        self.assertEqual(log.changes.get('deforestation_check_run'), 'clear')

    def test_override_fields_hidden_from_staff(self):
        """The disqualification override is manager-only — staff never see the
        fields on the farm form, so they can't set it."""
        from apps.suppliers.forms import FarmUpdateForm
        staff_form = FarmUpdateForm(company=self.company, user=self.staff, instance=self.farm)
        self.assertNotIn('land_cleared_after_cutoff', staff_form.fields)
        self.assertNotIn('land_cleared_after_cutoff_reason', staff_form.fields)
        manager_form = FarmUpdateForm(company=self.company, user=self.manager, instance=self.farm)
        self.assertIn('land_cleared_after_cutoff', manager_form.fields)
        self.assertIn('land_cleared_after_cutoff_reason', manager_form.fields)

    def test_manager_override_lets_flagged_farm_be_signed_off(self):
        """End-to-end: a flagged farm can't be signed off — but once a manager
        records a disqualification override, the evidence gate clears and the
        sign-off goes through."""
        DeforestationCheck.objects.create(
            farm=self.farm, company=self.company, risk_status='flagged',
            engine_status='complete',
            geometry_hash_at_assessment=self.farm.geometry_hash,
            assessed_at=timezone.now(),
        )
        self.client.force_login(self.manager)
        # Flagged → sign-off rejected.
        self.client.post(self._confirm_url())
        self.farm.refresh_from_db()
        self.assertFalse(self.farm.is_eudr_verified)
        # Manager records the override with a reason, then signs off.
        self.farm.land_cleared_after_cutoff = False
        self.farm.land_cleared_after_cutoff_reason = 'Clearing predates the cut-off — confirmed on the ground.'
        self.farm.save()
        r = self.client.post(self._confirm_url())
        self.assertEqual(r.status_code, 302)
        self.farm.refresh_from_db()
        self.assertTrue(self.farm.is_eudr_verified)


class BulkComplianceReadinessTests(TestCase):
    """
    Bulk Compliance Readiness sign-off — the bulk action must enforce the same
    evidence gate as the single-farm action, per farm, server-side. A stale
    browser tab (or a farm whose evidence drifted between page-load and submit)
    must NEVER produce a silent sign-off on an unready farm.
    """

    def setUp(self):
        self.company = Company.objects.create(
            name='Bulk Co', country='Nigeria', plan_tier='starter'
        )
        self.other_company = Company.objects.create(
            name='Bulk Other Co', country='Nigeria', plan_tier='starter'
        )
        self.supplier = Supplier.objects.create(
            company=self.company, name='Bulk Supplier', category='cooperative'
        )
        self.manager = CustomUser.objects.create_user(
            username='bulk_mgr', password='testpass',
            company=self.company, system_role='manager',
        )
        self.staff = CustomUser.objects.create_user(
            username='bulk_staff', password='testpass',
            company=self.company, system_role='staff',
        )
        self.url = reverse('suppliers:farm_bulk_signoff')

    def _farm(self, name='Bulk Farm', company=None, supplier=None, **kwargs):
        company = company or self.company
        supplier = supplier or self.supplier
        defaults = dict(
            name=name, country='Nigeria', commodity='Soy',
            geolocation=_POLYGON, deforestation_risk_status='low',
        )
        defaults.update(kwargs)
        return Farm.objects.create(company=company, supplier=supplier, **defaults)

    def _clear_check(self, farm):
        return DeforestationCheck.objects.create(
            farm=farm, company=farm.company, risk_status='clear',
            engine_status='complete',
            geometry_hash_at_assessment=farm.geometry_hash,
            assessed_at=timezone.now(),
        )

    def test_get_lists_awaiting_signoff_farms(self):
        ready_farm = self._farm(name='Ready One')
        self._clear_check(ready_farm)
        # Second farm with no evidence — has blockers, should NOT appear.
        self._farm(name='Blocked One')

        self.client.force_login(self.manager)
        r = self.client.get(self.url)

        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Ready One')
        self.assertNotContains(r, 'Blocked One')
        self.assertEqual(r.context['awaiting_count'], 1)
        self.assertEqual(r.context['expired_count'], 0)

    def test_get_lists_expired_farms_with_intact_evidence(self):
        farm = self._farm(name='Expired Renew')
        self._clear_check(farm)
        farm.is_eudr_verified = True
        farm.verified_by = self.manager
        farm.verified_date = datetime.date.today() - datetime.timedelta(days=400)
        farm.verification_expiry = datetime.date.today() - datetime.timedelta(days=35)
        farm.save()

        self.client.force_login(self.manager)
        r = self.client.get(self.url)

        self.assertEqual(r.context['expired_count'], 1)
        self.assertContains(r, 'Expired Renew')

    def test_get_excludes_disqualified_farms(self):
        farm = self._farm(name='Disqualified One')
        farm.land_cleared_after_cutoff = True
        farm.save()

        self.client.force_login(self.manager)
        r = self.client.get(self.url)

        self.assertNotContains(r, 'Disqualified One')
        self.assertEqual(r.context['eligible_count'], 0)

    def test_get_does_not_show_other_tenant_farms(self):
        other_supplier = Supplier.objects.create(
            company=self.other_company, name='Other Sup', category='cooperative'
        )
        other = self._farm(name='Other Tenant Farm', company=self.other_company, supplier=other_supplier)
        self._clear_check(other)

        self.client.force_login(self.manager)
        r = self.client.get(self.url)

        self.assertNotContains(r, 'Other Tenant Farm')

    def test_post_signs_off_selected_farms_and_writes_audit_log(self):
        from apps.audit.models import AuditLog
        f1 = self._farm(name='Sign One')
        f2 = self._farm(name='Sign Two')
        self._clear_check(f1)
        self._clear_check(f2)

        self.client.force_login(self.manager)
        r = self.client.post(self.url, {'farm_pks': [str(f1.pk), str(f2.pk)]})

        self.assertEqual(r.status_code, 200)
        f1.refresh_from_db()
        f2.refresh_from_db()
        self.assertTrue(f1.is_eudr_verified)
        self.assertTrue(f2.is_eudr_verified)
        self.assertEqual(f1.verified_by, self.manager)
        self.assertEqual(f2.verified_by, self.manager)

        # Per-farm audit logs (bulk action does NOT short-cut audit).
        f1_logs = AuditLog.objects.filter(model_name='Farm', object_repr__icontains='Sign One')
        f2_logs = AuditLog.objects.filter(model_name='Farm', object_repr__icontains='Sign Two')
        self.assertTrue(f1_logs.exists())
        self.assertTrue(f2_logs.exists())
        self.assertEqual(f1_logs.first().changes.get('compliance_readiness'), 'signed off (bulk)')

    def test_post_skips_farm_whose_evidence_drifted(self):
        """The killer test: between page-load and submit, evidence on one farm
        becomes invalid. The bulk POST must NOT sign off that farm."""
        ready = self._farm(name='Still Ready')
        drifted = self._farm(name='Drifted')
        self._clear_check(ready)
        self._clear_check(drifted)
        # Drift: a second flagged check is added on the drifted farm AFTER the
        # browser would have rendered the page.
        DeforestationCheck.objects.create(
            farm=drifted, company=self.company, risk_status='flagged',
            engine_status='complete',
            geometry_hash_at_assessment=drifted.geometry_hash,
            assessed_at=timezone.now(),
        )

        self.client.force_login(self.manager)
        self.client.post(self.url, {'farm_pks': [str(ready.pk), str(drifted.pk)]})

        ready.refresh_from_db()
        drifted.refresh_from_db()
        self.assertTrue(ready.is_eudr_verified)
        self.assertFalse(drifted.is_eudr_verified)

    def test_post_cannot_sign_off_other_tenant_farm(self):
        other_supplier = Supplier.objects.create(
            company=self.other_company, name='Other Sup', category='cooperative'
        )
        other = self._farm(name='Cross Tenant', company=self.other_company, supplier=other_supplier)
        self._clear_check(other)

        self.client.force_login(self.manager)
        self.client.post(self.url, {'farm_pks': [str(other.pk)]})

        other.refresh_from_db()
        self.assertFalse(other.is_eudr_verified)

    def test_staff_cannot_post_bulk_signoff(self):
        ready = self._farm(name='Staff Cannot')
        self._clear_check(ready)
        self.client.force_login(self.staff)
        r = self.client.post(self.url, {'farm_pks': [str(ready.pk)]})
        self.assertEqual(r.status_code, 403)
        ready.refresh_from_db()
        self.assertFalse(ready.is_eudr_verified)

    def test_staff_cannot_get_bulk_signoff(self):
        self.client.force_login(self.staff)
        r = self.client.get(self.url)
        self.assertEqual(r.status_code, 403)


class FarmerImportTests(TestCase):
    """
    Phase 1 of the farmer-import redesign: surface every transformation,
    loud-fail on unparseable values, write an auditable FarmerImportLog.

    Each test names the silent-drop path it guards against — these are
    behaviours the old importer would have hidden under a row count.
    """

    def setUp(self):
        from apps.suppliers.models import Farmer, FarmerImportLog
        self.Farmer = Farmer
        self.FarmerImportLog = FarmerImportLog
        self.company = Company.objects.create(
            name='Import Co', country='Nigeria', plan_tier='starter'
        )
        self.other = Company.objects.create(
            name='Other Co', country='Nigeria', plan_tier='starter'
        )
        self.staff = CustomUser.objects.create_user(
            username='import_staff', password='x',
            company=self.company, system_role='manager',
        )
        self.url = reverse('suppliers:farmer_import')

    def _upload(self, csv_text, filename='farmers.csv', encoding='utf-8'):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_login(self.staff)
        f = SimpleUploadedFile(filename, csv_text.encode(encoding), content_type='text/csv')
        return self.client.post(self.url, {'import_file': f})

    def _make_xlsx_bytes(self, rows):
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['first_name', 'last_name', 'gender', 'phone', 'village', 'lga', 'nin', 'crops', 'consent_date']
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _upload_xlsx(self, xlsx_bytes, filename='farmers.xlsx'):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_login(self.staff)
        ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        f = SimpleUploadedFile(filename, xlsx_bytes, content_type=ct)
        return self.client.post(self.url, {'import_file': f})

    # ── Validator unit tests ────────────────────────────────────────────────

    def test_validator_rejects_unknown_gender(self):
        """Old importer silently blanked unrecognised gender. Must now reject."""
        from apps.suppliers.farmer_import import validate_farmer_row
        result = validate_farmer_row(
            {'first_name': 'Amina', 'gender': 'X'}, row_num=2,
        )
        self.assertTrue(result['errors'])
        self.assertEqual(result['errors'][0]['field'], 'gender')

    def test_validator_rejects_unparseable_consent_date(self):
        """Old importer silently dropped bad dates. Must now reject with raw echoed."""
        from apps.suppliers.farmer_import import validate_farmer_row
        result = validate_farmer_row(
            {'first_name': 'A', 'consent_date': 'last tuesday'}, row_num=2,
        )
        self.assertTrue(result['errors'])
        self.assertEqual(result['errors'][0]['field'], 'consent_date')
        self.assertIn("'last tuesday'", result['errors'][0]['reason'])

    def test_validator_accepts_multiple_date_formats(self):
        from apps.suppliers.farmer_import import validate_farmer_row
        for raw in ('2026-04-01', '01/04/2026', '01-04-2026', '2026/04/01'):
            r = validate_farmer_row({'first_name': 'A', 'consent_date': raw}, 2)
            self.assertFalse(r['errors'], f"format {raw} should parse")
            self.assertEqual(r['fields']['consent_date'], datetime.date(2026, 4, 1))

    def test_validator_surfaces_phone_normalisation(self):
        """Old importer silently normalised on save. Must surface as transformation."""
        from apps.suppliers.farmer_import import validate_farmer_row
        r = validate_farmer_row({'first_name': 'A', 'phone': '08012345678'}, 2)
        self.assertEqual(r['fields']['phone'], '+2348012345678')
        self.assertEqual(len(r['transformations']), 1)
        self.assertEqual(r['transformations'][0]['field'], 'phone')
        self.assertEqual(r['transformations'][0]['from'], '08012345678')
        self.assertEqual(r['transformations'][0]['to'],   '+2348012345678')

    def test_validator_surfaces_nin_cleanup(self):
        from apps.suppliers.farmer_import import validate_farmer_row
        r = validate_farmer_row({'first_name': 'A', 'nin': '123-456-78901'}, 2)
        self.assertEqual(r['fields']['nin'], '12345678901')
        nin_tx = [t for t in r['transformations'] if t['field'] == 'nin']
        self.assertEqual(len(nin_tx), 1)

    def test_validator_warns_on_short_nin(self):
        """NIN with wrong length should warn, not block — co-ops have imperfect records."""
        from apps.suppliers.farmer_import import validate_farmer_row
        r = validate_farmer_row({'first_name': 'A', 'nin': '12345'}, 2)
        self.assertFalse(r['errors'])
        self.assertTrue(any(w['field'] == 'nin' for w in r['warnings']))

    def test_validator_first_name_required(self):
        from apps.suppliers.farmer_import import validate_farmer_row
        r = validate_farmer_row({'first_name': ''}, 2)
        self.assertTrue(r['errors'])
        self.assertEqual(r['errors'][0]['field'], 'first_name')

    # ── End-to-end view tests ───────────────────────────────────────────────

    def _csv_header(self):
        return 'first_name,last_name,gender,phone,village,lga,nin,crops,consent_date\n'

    def test_post_creates_farmer_and_writes_log(self):
        csv_text = self._csv_header() + 'Amina,Musa,F,08012345678,Shendam,Shendam,12345678901,Soybeans,2026-04-01\n'
        r = self._upload(csv_text)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(self.Farmer.objects.filter(company=self.company).count(), 1)
        # Log written, scoped to tenant
        logs = self.FarmerImportLog.objects.filter(company=self.company)
        self.assertEqual(logs.count(), 1)
        self.assertEqual(logs.first().created, 1)
        self.assertEqual(logs.first().uploaded_by, self.staff)

    def test_post_rejects_bad_gender_loudly(self):
        csv_text = self._csv_header() + 'Amina,Musa,X,08012345678,Shendam,Shendam,12345678901,Soybeans,2026-04-01\n'
        r = self._upload(csv_text)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(self.Farmer.objects.filter(company=self.company).count(), 0)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.errors, 1)
        self.assertIn("'X'", log.error_detail[0]['error_reason'])

    def test_post_rejects_bad_consent_date_loudly(self):
        csv_text = self._csv_header() + 'Amina,Musa,F,08012345678,Shendam,Shendam,12345678901,Soybeans,banana\n'
        r = self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.errors, 1)
        self.assertIn('banana', log.error_detail[0]['error_reason'])
        # Raw row is echoed for the error CSV download
        self.assertEqual(log.error_detail[0]['consent_date'], 'banana')

    def test_post_records_phone_normalisation_as_transformation(self):
        csv_text = self._csv_header() + 'Amina,Musa,F,08012345678,Shendam,Shendam,,,\n'
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertEqual(log.auto_corrected, 1)
        self.assertEqual(len(log.transformation_log), 1)
        self.assertEqual(log.transformation_log[0]['field'], 'phone')

    def test_post_records_nin_length_warning_without_blocking(self):
        csv_text = self._csv_header() + 'Amina,Musa,F,,Shendam,Shendam,12345,,\n'
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertEqual(log.warning_count, 1)
        self.assertEqual(len(log.warning_detail), 1)
        self.assertEqual(log.warning_detail[0]['field'], 'nin')

    def test_post_dedupes_by_nin(self):
        self.Farmer.objects.create(
            company=self.company, first_name='Existing', nin='12345678901',
        )
        csv_text = self._csv_header() + 'Amina,Musa,F,,,, 12345678901 ,,\n'
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 0)
        self.assertEqual(log.duplicates, 1)

    def test_post_dedupes_by_name_village_lga(self):
        self.Farmer.objects.create(
            company=self.company,
            first_name='Amina', last_name='Musa', village='Shendam', lga='Shendam',
        )
        csv_text = self._csv_header() + 'Amina,Musa,F,,Shendam,Shendam,,,\n'
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 0)
        self.assertEqual(log.duplicates, 1)

    def test_post_intra_batch_dedupe(self):
        """Same NIN appearing twice in the same upload must dedup, not create twice."""
        csv_text = self._csv_header() + (
            'Amina,Musa,F,,Shendam,Shendam,12345678901,,\n'
            'Amina,Musa,F,,Shendam,Shendam,12345678901,,\n'
        )
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertEqual(log.duplicates, 1)

    def test_encoding_fallback_warns_on_cp1252(self):
        """Files saved by Excel without UTF-8 should import but emit a fallback warning."""
        # "Børje" — the ø encodes differently in CP1252 (0xF8) vs UTF-8 (0xC3 0xB8)
        # so the UTF-8 decode fails and the chain falls back to cp1252.
        csv_text = self._csv_header() + 'Børje,Musa,F,,Shendam,Shendam,,,\n'
        r = self._upload(csv_text, encoding='cp1252')
        self.assertEqual(r.status_code, 200)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertTrue(any(
            w.get('field') == 'file_encoding' for w in log.warning_detail
        ))

    def test_file_size_limit_returns_friendly_error(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from apps.suppliers.farmer_import import MAX_FARMER_IMPORT_BYTES
        self.client.force_login(self.staff)
        big = SimpleUploadedFile(
            'big.csv',
            b'x' * (MAX_FARMER_IMPORT_BYTES + 1),
            content_type='text/csv',
        )
        r = self.client.post(self.url, {'import_file': big})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'larger than')
        self.assertEqual(self.FarmerImportLog.objects.count(), 0)

    def test_errors_download_scoped_to_tenant(self):
        # Create a log on the OTHER company and confirm our staff can't read it
        other_log = self.FarmerImportLog.objects.create(
            company=self.other, error_detail=[{'first_name': 'leaked'}],
        )
        self.client.force_login(self.staff)
        r = self.client.get(reverse('suppliers:farmer_import_errors', args=[other_log.pk]))
        self.assertEqual(r.status_code, 404)

    def test_errors_download_returns_csv(self):
        csv_text = self._csv_header() + 'Amina,Musa,X,,,,,,\n'
        self._upload(csv_text)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        r = self.client.get(reverse('suppliers:farmer_import_errors', args=[log.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'text/csv')
        body = r.content.decode()
        self.assertIn('error_reason', body)
        self.assertIn('Amina', body)

    # ── Phase 2: XLSX upload tests ──────────────────────────────────────────

    def test_xlsx_creates_farmer_same_as_csv(self):
        """XLSX upload produces the same result as the equivalent CSV row."""
        xlsx = self._make_xlsx_bytes([{
            'first_name': 'Amina', 'last_name': 'Musa', 'gender': 'F',
            'phone': '08012345678', 'village': 'Shendam', 'lga': 'Shendam',
            'nin': '12345678901', 'crops': 'Soybeans', 'consent_date': '2026-04-01',
        }])
        r = self._upload_xlsx(xlsx)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(self.Farmer.objects.filter(company=self.company).count(), 1)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertEqual(log.errors, 0)

    def test_xlsx_rejects_bad_gender_loudly(self):
        """XLSX rows go through validate_farmer_row — bad gender must still reject."""
        xlsx = self._make_xlsx_bytes([{'first_name': 'Amina', 'gender': 'X'}])
        self._upload_xlsx(xlsx)
        self.assertEqual(self.Farmer.objects.filter(company=self.company).count(), 0)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.errors, 1)
        self.assertIn("'X'", log.error_detail[0]['error_reason'])

    def test_xlsx_records_phone_normalisation_as_transformation(self):
        """Phone normalisation is surfaced as a transformation for XLSX rows too."""
        xlsx = self._make_xlsx_bytes([{
            'first_name': 'Amina', 'phone': '08012345678',
        }])
        self._upload_xlsx(xlsx)
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertEqual(log.auto_corrected, 1)
        self.assertEqual(log.transformation_log[0]['field'], 'phone')

    def test_xlsx_multi_sheet_warns_and_imports_first_sheet_only(self):
        """Multiple sheets: import first sheet, attach a sheet_selection warning."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Farmers'
        ws1.append(['first_name', 'last_name', 'gender', 'phone', 'village', 'lga', 'nin', 'crops', 'consent_date'])
        ws1.append(['Amina', 'Musa', 'F', '', '', '', '', '', ''])
        ws2 = wb.create_sheet('Ignore')
        ws2.append(['do', 'not', 'import', '', '', '', '', '', ''])
        ws2.append(['Extra', 'Row', '', '', '', '', '', '', ''])
        buf = io.BytesIO()
        wb.save(buf)
        self._upload_xlsx(buf.getvalue())
        log = self.FarmerImportLog.objects.filter(company=self.company).first()
        self.assertEqual(log.created, 1)
        self.assertTrue(any(w.get('field') == 'sheet_selection' for w in log.warning_detail))

    def test_xlsx_magic_bytes_with_csv_extension_is_rejected(self):
        """An XLSX binary uploaded with .csv extension should return a friendly error."""
        xlsx = self._make_xlsx_bytes([{'first_name': 'Amina'}])
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_login(self.staff)
        f = SimpleUploadedFile('farmers.csv', xlsx, content_type='text/csv')
        r = self.client.post(self.url, {'import_file': f})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Excel')
        self.assertEqual(self.Farmer.objects.filter(company=self.company).count(), 0)
        self.assertEqual(self.FarmerImportLog.objects.count(), 0)

    def test_unsupported_extension_returns_friendly_error(self):
        """Files with unrecognised extensions (.pdf, .doc, etc.) are rejected."""
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_login(self.staff)
        f = SimpleUploadedFile('farmers.pdf', b'%PDF-1.4 fake', content_type='application/pdf')
        r = self.client.post(self.url, {'import_file': f})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Unsupported')
        self.assertEqual(self.FarmerImportLog.objects.count(), 0)
